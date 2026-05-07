"""Export data/law_index.json to an Excel workbook with hyperlinks.

Output: output/law_index_summary.xlsx
- Sheet 1 "法規總覽": one row per law with category, name, authority, links
- Sheet 2 "分類統計": category counts vs. expected
- Sheet 3 "常見條文": flattened common_articles per law

Re-run this script whenever data/law_index.json is updated.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "law_index.json"
DST = ROOT / "output" / "law_index_summary.xlsx"


def _hyperlink(cell, url: str | None) -> None:
    if not url:
        cell.value = ""
        return
    cell.value = url
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            longest_line = max((len(line) for line in value.splitlines()), default=0)
            length = max(length, longest_line)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, max_width)


def main() -> None:
    data = json.loads(SRC.read_text(encoding="utf-8"))
    categories = {c["code"]: c for c in data["categories"]}
    laws = data["laws"]

    wb = Workbook()

    overview = wb.active
    overview.title = "法規總覽"
    headers = [
        "編號",
        "分類代碼",
        "分類名稱",
        "法規全名",
        "簡稱",
        "主管機關",
        "全國法規資料庫連結",
        "證交所 selaw 連結",
        "條文 URL 範本",
        "常用條文數",
        "搜尋關鍵字",
    ]
    overview.append(headers)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in overview[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")
    overview.freeze_panes = "A2"

    for law in laws:
        cat_code = law.get("category", "")
        cat_name = categories.get(cat_code, {}).get("name", "")
        row_idx = overview.max_row + 1
        overview.cell(row=row_idx, column=1, value=law.get("id"))
        overview.cell(row=row_idx, column=2, value=cat_code)
        overview.cell(row=row_idx, column=3, value=cat_name)
        overview.cell(row=row_idx, column=4, value=law.get("name"))
        overview.cell(row=row_idx, column=5, value=law.get("abbreviation"))
        overview.cell(row=row_idx, column=6, value=law.get("issuing_authority"))
        _hyperlink(overview.cell(row=row_idx, column=7), law.get("primary_url"))
        _hyperlink(overview.cell(row=row_idx, column=8), law.get("selaw_url"))
        overview.cell(row=row_idx, column=9, value=law.get("article_url_template"))
        overview.cell(row=row_idx, column=10, value=len(law.get("common_articles") or []))
        overview.cell(
            row=row_idx,
            column=11,
            value="、".join(law.get("search_keywords") or []),
        )

    _autosize(overview)

    stats = wb.create_sheet("分類統計")
    stats.append(["分類代碼", "分類名稱", "預期法規數", "實際法規數", "差異"])
    for cell in stats[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")

    actual_counts: dict[str, int] = {}
    for law in laws:
        actual_counts[law["category"]] = actual_counts.get(law["category"], 0) + 1

    for cat in data["categories"]:
        code = cat["code"]
        actual = actual_counts.get(code, 0)
        stats.append(
            [
                code,
                cat["name"],
                cat["expected_count"],
                actual,
                actual - cat["expected_count"],
            ]
        )
    stats.append(
        [
            "合計",
            "",
            sum(c["expected_count"] for c in data["categories"]),
            sum(actual_counts.values()),
            sum(actual_counts.values())
            - sum(c["expected_count"] for c in data["categories"]),
        ]
    )
    _autosize(stats)

    articles = wb.create_sheet("常見條文")
    articles.append(["法規編號", "法規全名", "條號", "主題", "條文 URL"])
    for cell in articles[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for law in laws:
        template = law.get("article_url_template") or ""
        for art in law.get("common_articles") or []:
            no = art.get("no", "")
            row_idx = articles.max_row + 1
            articles.cell(row=row_idx, column=1, value=law.get("id"))
            articles.cell(row=row_idx, column=2, value=law.get("name"))
            articles.cell(row=row_idx, column=3, value=no)
            articles.cell(row=row_idx, column=4, value=art.get("topic"))
            url = template.replace("{article_no}", str(no)) if template and no else None
            _hyperlink(articles.cell(row=row_idx, column=5), url)

    _autosize(articles)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)
    print(f"Wrote {DST.relative_to(ROOT)} ({len(laws)} laws)")


if __name__ == "__main__":
    main()
